import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_weekly_excel(leads, output_path):
    wb = Workbook()
    ws = wb.active

    # Week label
    from datetime import datetime, timedelta
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    week_num = today.isocalendar()[1]
    ws.title = f"Week {week_num} — {monday.strftime('%d-%m')} to {sunday.strftime('%d-%m')}"

    # Colors
    YELLOW   = "FFD700"
    ORANGE   = "FFA040"
    PINK     = "FFB6C1"
    GRAY     = "D3D3D3"
    GREEN_BG = "E8F5E9"
    PINK_BG  = "FCE4EC"
    WHITE    = "FFFFFF"

    STATUS_COLOR = {
        "interested":       GREEN_BG,
        "waiting":          GREEN_BG,
        "follow_up_needed": GREEN_BG,
        "no_budget":        PINK_BG,
        "closed_lost":      PINK_BG,
    }

    def hdr_style(hex_color):
        return {
            "fill": PatternFill("solid", fgColor=hex_color),
            "font": Font(bold=True, size=11, name="Arial"),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border": Border(
                top=Side(style="thin"), bottom=Side(style="thin"),
                left=Side(style="thin"), right=Side(style="thin")
            )
        }

    def apply(cell, styles):
        for k, v in styles.items():
            setattr(cell, k, v)

    # Column widths
    col_widths = [6, 20, 28, 30, 30, 45, 35, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1: Main headers
    headers = ["STT", "Name", "Website", "Sources", "Reach out", "Research", "Note", "Crosscheck", "CEO Check"]
    colors =  [YELLOW, YELLOW, YELLOW, YELLOW, ORANGE, YELLOW, PINK, GRAY, GRAY]
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    for col, (h, clr) in enumerate(zip(headers, colors), 1):
        cell = ws.cell(row=1, column=col, value=h)
        apply(cell, hdr_style(clr))

    # Row 2: Sub-header "TG" under Reach out (col E)
    sub = ws.cell(row=2, column=5, value="TG")
    apply(sub, hdr_style(ORANGE))

    # Merge rows 1-2 for all columns except E
    for col in [1, 2, 3, 4, 6, 7, 8, 9]:
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # Data rows
    def cell_style(hex_bg):
        return {
            "fill": PatternFill("solid", fgColor=hex_bg),
            "alignment": Alignment(vertical="top", wrap_text=True),
            "border": Border(
                top=Side(style="thin"), bottom=Side(style="thin"),
                left=Side(style="thin"), right=Side(style="thin")
            )
        }

    for idx, lead in enumerate(leads, 1):
        row = idx + 2
        ws.row_dimensions[row].height = 90
        bg = STATUS_COLOR.get(lead.get("status", ""), WHITE)
        st = cell_style(bg)

        values = [
            idx,
            lead.get("name", ""),
            lead.get("website", ""),
            lead.get("sources", ""),
            lead.get("telegram_username", ""),
            lead.get("research", ""),
            lead.get("note", ""),
            "",  # Crosscheck
            "",  # CEO Check
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            apply(cell, st)
            cell.font = Font(name="Arial", size=10)

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")

if __name__ == "__main__":
    leads = json.loads(sys.argv[1])
    output = sys.argv[2]
    generate_weekly_excel(leads, output)
